#!/usr/bin/env python3
"""inventario_carpeta.py — descubre y clasifica los documentos de un expediente del Impuesto sobre
Sociedades dentro de una carpeta local. DETERMINISTA: el parser hace el trabajo (no el LLM), para no
disparar tokens leyendo PDFs. Asigna a cada fichero su PAPEL (sys / prev200 / modelo200_pdf /
datos_fiscales / liquidacion / ccaa / salida / otro), detecta el formato del Sumas y Saldos (cabecera
ES/EN, multipestaña) y devuelve un manifiesto de intake para iterar con el usuario antes de correr el motor.

Uso:  python3 inventario_carpeta.py "<ruta-de-la-carpeta>"
PII:  los nombres de fichero son locales; en el chat resume en CODENAME, sin NIF/importes.
"""
import sys, os, json, glob, re, unicodedata

SYS_EXT = (".xlsx", ".xls", ".csv")


def _norm_text(s):
    return unicodedata.normalize("NFC", str(s or "")).lower()


def _kw(name, *ks):
    n = _norm_text(name)
    return any(_norm_text(k) in n for k in ks)


def _rel(folder, path):
    return os.path.relpath(path, folder)


def _sniff_200(path):
    try:
        with open(path, "rb") as f:
            head = f.read(64).lstrip()
        return head[:5] == b"<T200"
    except Exception:
        return False


def _num(v):
    s = str(v if v is not None else "").strip()
    if not s:
        return 0.0
    s = s.replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(re.sub(r"[^0-9.\-]", "", s) or 0)
    except ValueError:
        return 0.0


def _sniff_sys_xlsx(path):
    """Metadatos ligeros del SyS. Best-effort; degrada si falta openpyxl o el libro es ilegible."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sh in wb.sheetnames:
                ws = wb[sh]
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                    cells = [str(c).strip().lower() for c in row if c is not None]
                    acc = any(c.startswith(("cuenta", "codig", "códig", "account")) for c in cells)
                    desc = any(("descrip" in c) or ("concepto" in c) or ("description" in c) for c in cells)
                    if acc and desc:
                        en = any(c.startswith("account") or "description" in c or "debit" in c or "credit" in c for c in cells)
                        headers = [str(c).strip().lower() if c is not None else "" for c in row]

                        def idx(*names):
                            for i, h in enumerate(headers):
                                if any(n in h for n in names):
                                    return i
                            return None

                        i_cuenta = idx("cuenta", "account", "código", "codigo")
                        i_si = idx("saldo inicial", "saldo ant", "opening", "initial")
                        i_sf = idx("saldo final", "closing", "final")
                        n_rows = n_si_nonzero = n_pn = n_pn_si_nonzero = n_pn_sf_nonzero = 0
                        for data in ws.iter_rows(min_row=row_idx + 1, values_only=True):
                            if i_cuenta is None or i_cuenta >= len(data):
                                continue
                            cuenta = str(data[i_cuenta] or "").strip().split(".")[0]
                            if not cuenta:
                                continue
                            n_rows += 1
                            si = _num(data[i_si]) if i_si is not None and i_si < len(data) else 0.0
                            sf = _num(data[i_sf]) if i_sf is not None and i_sf < len(data) else 0.0
                            if abs(si) > 0.005:
                                n_si_nonzero += 1
                            if re.match(r"^1\d{2,3}$", cuenta):
                                n_pn += 1
                                if abs(si) > 0.005:
                                    n_pn_si_nonzero += 1
                                if abs(sf) > 0.005:
                                    n_pn_sf_nonzero += 1
                        return {
                            "legible": True,
                            "idioma": "EN" if en else "ES",
                            "hojas": len(wb.sheetnames),
                            "hoja_tb": sh,
                            "filas": n_rows,
                            "tiene_saldo_inicial": i_si is not None,
                            "saldo_inicial_no_cero": n_si_nonzero,
                            "cuentas_pn": n_pn,
                            "pn_saldo_inicial_no_cero": n_pn_si_nonzero,
                            "pn_saldo_final_no_cero": n_pn_sf_nonzero,
                        }
            return {"legible": False, "idioma": None, "hojas": len(wb.sheetnames), "hoja_tb": None}
        finally:
            wb.close()
    except Exception:
        return {"legible": None, "idioma": None, "hojas": None, "hoja_tb": None}


def scan(folder):
    files, bypath = [], {}
    salida_files = []
    for p in sorted(glob.glob(os.path.join(folder, "**", "*"), recursive=True)):
        b = os.path.basename(p)
        if os.path.isdir(p) or b.startswith(("~$", ".")):
            continue
        if (os.sep + "salida" + os.sep) in p:
            salida_files.append(p)
            continue
        files.append(p); bypath[b] = p
    roles = {
        "sys": [],
        "prev200": [],
        "modelo200_pdf": [],
        "datos_fiscales": [],
        "liquidacion": [],
        "ccaa": [],
        "salida": [os.path.basename(p) for p in salida_files],
        "otro": [],
    }
    paths = {k: [] for k in roles}
    paths["salida"] = [_rel(folder, p) for p in salida_files]
    notes = []
    for p in files:
        b = os.path.basename(p); ext = os.path.splitext(b.lower())[1]
        if ext == ".200" or _sniff_200(p):
            roles["prev200"].append(b)
            paths["prev200"].append(_rel(folder, p))
        elif ext in SYS_EXT and _kw(b, "sumas", "saldo", "balance", "sys", "trial"):
            roles["sys"].append(b)
            paths["sys"].append(_rel(folder, p))
        elif ext in SYS_EXT and _kw(b, "gis", "liquidacion", "liquidación", "estimacion", "estimación"):
            roles["liquidacion"].append(b)
            paths["liquidacion"].append(_rel(folder, p))
        elif ext == ".pdf" and _kw(b, "datos fiscales", "datos_fiscales", "fiscales",
                                   "informac. fiscal", "informacion fiscal", "información fiscal"):
            roles["datos_fiscales"].append(b)
            paths["datos_fiscales"].append(_rel(folder, p))
        elif ext == ".pdf" and _kw(b, "modelo 200", "modelo200", "justificante", "autoliquid", "declaracion"):
            roles["modelo200_pdf"].append(b)
            paths["modelo200_pdf"].append(_rel(folder, p))
        elif ext in (".pdf", ".docx") and _kw(b, "cuentas anuales", "ccaa", "cuentas_anuales", "memoria",
                                               "annual accounts", "fy25 accounts", "financial statements"):
            roles["ccaa"].append(b)
            paths["ccaa"].append(_rel(folder, p))
        elif ext in SYS_EXT:
            roles["otro"].append(b + "  (Excel/CSV — ¿contabilidad?)")
            paths["otro"].append(_rel(folder, p))
        else:
            roles["otro"].append(b)
            paths["otro"].append(_rel(folder, p))
    if not roles["sys"]:
        xl = [b for b in bypath if os.path.splitext(b.lower())[1] in (".xlsx", ".xls")]
        if len(xl) == 1:
            roles["sys"].append(xl[0])
            paths["sys"].append(_rel(folder, bypath[xl[0]]))
            notes.append("SyS asumido por ser el único Excel de la carpeta.")
    sysinfo = None
    if roles["sys"]:
        sp = bypath.get(roles["sys"][0])
        if sp and sp.lower().endswith(".xlsx"):
            sysinfo = _sniff_sys_xlsx(sp)
            if sysinfo.get("hojas") and sysinfo.get("hojas") > 1:
                notes.append(f"El SyS es un libro multipestaña ({sysinfo.get('hojas')} hojas); el TB parece estar en la hoja «{sysinfo.get('hoja_tb')}».")
            if sysinfo.get("idioma") == "EN":
                notes.append("Cabecera del SyS en INGLÉS — el motor ya lo soporta (v1.7.0+).")
    faltan = [r for r in ("sys",) if not roles[r]]
    reco = [r for r in ("prev200", "datos_fiscales") if not roles[r]]
    preguntas = []
    bloqueos_semanticos = []
    if len(roles["sys"]) > 1:
        preguntas.append("Hay varios posibles SyS. Confirma cuál debe usar el motor.")
    # APERTURA del ECPN: el SyS debe traer saldo de apertura (N-1). Si no, el ECPN sale con apertura 0 y la AEAT
    # lo recalcula y rechaza (E25400632/E25400645). PERO el motor sabe DERIVAR la apertura de la columna N-1 del
    # Balance de la CCAA (engines.extraer_texto_ccaa → ecpn_sys.proyectar_desde_balance). Por eso: si hay CCAA, es
    # AVISO (el motor la deriva; revisar ECPN en Fase 2); si NO hay CCAA, es BLOQUEO (no hay de dónde sacarla).
    apertura_aviso = None
    if sysinfo and sysinfo.get("legible"):
        if not sysinfo.get("tiene_saldo_inicial"):
            apertura_aviso = "El SyS no trae columna de Saldo Inicial/apertura (el ECPN necesita apertura N-1)."
        elif sysinfo.get("filas", 0) and sysinfo.get("saldo_inicial_no_cero", 0) == 0:
            apertura_aviso = "El SyS trae Saldo Inicial pero está TODO A CERO (export YTD: no es apertura real)."
        elif sysinfo.get("cuentas_pn", 0) and sysinfo.get("pn_saldo_inicial_no_cero", 0) == 0 and sysinfo.get("pn_saldo_final_no_cero", 0):
            apertura_aviso = "Las cuentas de patrimonio neto tienen cierre pero apertura cero (el ECPN no reconciliaría)."
    if apertura_aviso:
        if roles["ccaa"]:
            preguntas.append(apertura_aviso + " HAY CCAA: el motor tomará la apertura del ECPN de su columna N-1 "
                             "(revisar el cuadre del ECPN en Fase 2). Mejor aún: regenerar el SyS con apertura N-1 "
                             "real y PN acumulado (skill ccaa-a-sys-a3).")
        else:
            bloqueos_semanticos.append(apertura_aviso + " NO hay CCAA para derivar la apertura: regenera el SyS con "
                                       "apertura desde la columna N-1 del balance anterior, o aporta la CCAA.")
    if not roles["prev200"] and roles["modelo200_pdf"]:
        preguntas.append("Solo hay PDF del Modelo 200 N-1. Si tienes el `.200` N-1, aportarlo mejora la precarga determinista; si no, se usará el PDF como fuente asistida.")
    if not roles["prev200"] and not roles["modelo200_pdf"] and not roles["datos_fiscales"]:
        preguntas.append("Falta fuente N-1/datos fiscales para formales. El motor puede arrancar con SyS, pero la importabilidad requerirá completar identificativos/formales.")
    if roles["liquidacion"]:
        preguntas.append("Hay fichero GIS/liquidación. Confirmar si se usa solo como referencia post-import o como input fiscal validado por el abogado.")
    if roles["salida"]:
        preguntas.append("Hay `.200`/salidas generadas previamente en `salida/`. Se ignoran para el intake salvo que quieras diagnosticarlas como intento anterior.")
    for b in bloqueos_semanticos:
        preguntas.append(b)
    calidad = "bloqueado" if (faltan or bloqueos_semanticos) else (
        "completo" if roles["datos_fiscales"] and (roles["prev200"] or roles["modelo200_pdf"]) and roles["ccaa"]
        else "mínimo"
    )
    manifest = {
        "codename": os.path.basename(os.path.abspath(folder)),
        "folder": folder,
        "paths": paths,
        "seleccion_motor": {
            "sys": paths["sys"][0] if paths["sys"] else None,
            "prev200": paths["prev200"][0] if paths["prev200"] else None,
            "modelo200_pdf": paths["modelo200_pdf"][0] if paths["modelo200_pdf"] else None,
            "datos_fiscales": paths["datos_fiscales"][0] if paths["datos_fiscales"] else None,
            "ccaa": paths["ccaa"][0] if paths["ccaa"] else None,
            "liquidacion": paths["liquidacion"][0] if paths["liquidacion"] else None,
        },
        "listo_para_motor": not faltan and not bloqueos_semanticos,
        "calidad_intake": calidad,
        "faltan_requeridos": faltan,
        "faltan_recomendados": reco,
        "bloqueos_semanticos": bloqueos_semanticos,
        "preguntas_usuario": preguntas,
    }
    return roles, sysinfo, notes, faltan, reco, preguntas, manifest


def main():
    if len(sys.argv) < 2:
        sys.exit('Uso: python3 inventario_carpeta.py "<ruta-de-la-carpeta>"')
    folder = os.path.abspath(sys.argv[1])
    if not os.path.isdir(folder):
        sys.exit(f"ERROR: no existe la carpeta {folder}")
    roles, sysinfo, notes, faltan, reco, preguntas, manifest = scan(folder)
    LB = {"sys": "Sumas y Saldos (REQUERIDO)", "prev200": ".200 año anterior",
          "modelo200_pdf": "Modelo 200 / justificante N-1 (PDF)",
          "datos_fiscales": "Datos fiscales (PDF)", "liquidacion": "GIS / liquidación (Excel)",
          "ccaa": "Cuentas anuales (PDF/DOCX)", "salida": "Salidas previas", "otro": "Otros (revisar)"}
    print(f"INVENTARIO — {folder}")
    for r in ("sys", "prev200", "modelo200_pdf", "datos_fiscales", "liquidacion", "ccaa", "salida", "otro"):
        if roles[r]:
            print(f"  • {LB[r]}: " + " · ".join(roles[r]))
    if sysinfo and sysinfo.get("idioma"):
        print(f"  · SyS: idioma={sysinfo['idioma']} · hojas={sysinfo['hojas']} · hoja_TB={sysinfo['hoja_tb']}")
    for n in notes:
        print("  ⚠ " + n)
    print("  FALTA (requerido): " + (", ".join(LB[r] for r in faltan) if faltan else "nada"))
    if reco:
        print("  Recomendado (mejora la precarga): " + ", ".join(LB[r] for r in reco))
    if preguntas:
        print("  Preguntas al usuario:")
        for q in preguntas:
            print("    - " + q)
    print("  CALIDAD INTAKE: " + manifest["calidad_intake"])
    print("  LISTO PARA MOTOR: " + ("sí" if manifest["listo_para_motor"] else "no"))
    print("JSON " + json.dumps({"roles": roles, "sys": sysinfo, "faltan": faltan,
                                "recomendado": reco, "manifest": manifest}, ensure_ascii=False))


if __name__ == "__main__":
    main()
